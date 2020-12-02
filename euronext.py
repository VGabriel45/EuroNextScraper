import time
import openpyxl
import os
from bs4 import BeautifulSoup
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from openpyxl import Workbook
from webdriver_manager.chrome import ChromeDriverManager
print(" _______________________________")
print("|                               |")
print("|           WELCOME             |")
print("|   project Euronext NOS data   |")
print("|_______________________________|\n")

print("Loading Libraries...")


filenames = ['Belgium1.xlsx', 'France1.xlsx',
             'Ireland1.xlsx', 'Netherlands1.xlsx', 'Portugal1.xlsx']
files = list()

for n in range(5):
    if os.path.exists(filenames[n]):
        files.append(filenames[n])

if len(files) > 0:

    print("\nBelow files are in the current folder:\n")
    print("ID   Filename")
    print("--   --------")
    for nf in range(len(files)):
        print(str(nf) + ":   " + str(files[nf]))

    file = input("\nType the ID of file you want to scrape: ")

    if int(file) >= len(files):
        print("ID "+file+" does not exist.")
    else:
        line = int(input("From which line you want to scrape the data: "))

        print("\n " + "_" * (27 + len(str(line)) + len(files[int(file)])))
        print("|" + " " * (27 + len(str(line)) + len(files[int(file)])) + "|")
        print("|   SCRAPING " + files[int(file)] +
              " from line: " + str(line) + "   |")
        print("|" + "_" * (27 + len(str(line)) + len(files[int(file)])) + "|")

        print("\nStarting Chrome...")
        path = os.getcwd()
        full_path = path+str('/chromedriver')
        options = Options()
        # options.headless = True
        options.add_argument('--log-level=2')
        options.add_argument('--disable-gpu')
        driver = webdriver.Chrome(ChromeDriverManager().install())

        book = openpyxl.load_workbook("France1.xlsx")
        sheet = book.active
        row_count = sheet.max_row

        count = 0

        while line < row_count+1:
            try:
                print("\nProcessing " +
                      files[int(file)] + " line: " + str(line))
                url = sheet.cell(line, 42).value
                if url is None:
                    pass
                elif str(url).startswith("https://live.euronext.com/"):
                    driver.get(url)

                    wait = WebDriverWait(driver, 10)
                    shares = wait.until(EC.visibility_of_element_located(
                        (By.CSS_SELECTOR, "#content > section > div > div:nth-child(6) > div.card-body > div > table > tbody > tr:nth-child(4) > td:nth-child(2) > strong")))
                    page_source = driver.page_source
                    try:
                        soup = BeautifulSoup(page_source, 'lxml')
                        shares_outs = soup.find(
                            'td', text='Shares outstanding').find_next('td').text.strip()

                        f_shares = shares_outs.replace(",", "")
                        print("Shares Outstanding - " +
                              sheet.cell(line, 1).value + ": " + f_shares)
                        sheet.cell(line, 28).value = f_shares
                        line += 1
                        count = 0
                    except Exception:
                        print("Exception has been thrown. Moving to next line")
                        line += 1
                else:
                    print("URL is not right. Moving to the next line.")
                    line += 1

            except TimeoutException:
                if count == 0:
                    print("\n-Exception has been thrown. Pausing for 1 minute.-\n")
                    time.sleep(60)
                    count += 1
                else:
                    print(
                        "\n-Exception has been thrown for the 2nd time. Moving to next line-\n")
                    count = 0
                    line += 1

        book.save(files[int(file)])
        print("\n" + files[int(file)] + " has been saved.")

        driver.close()
        print("\n __________")
        print("|          |")
        print("|   DONE   |")
        print("|__________|")

else:
    print("No files found")
